from openpyxl import load_workbook, Workbook

from database.database import open_database, geef_chocoladeletters
from familielid import Familielid

if __name__ == '__main__':
    werkboek: Workbook = load_workbook("bestanden/Familiebestand.xlsx")
    familie_werkblad = werkboek["Familie"]
    cel1 = familie_werkblad.cell(1, 1)
    print(f'In de eerste cel staat: "{cel1.value}".')

    familieleden = list()
    for rijnummer, cellen in enumerate(familie_werkblad.iter_rows(min_row=2, max_col=3), start=2):
        # Lees de eerste drie kolommen
        naam = familie_werkblad.cell(rijnummer, 1).value
        familieverband = familie_werkblad.cell(rijnummer, 2).value
        geboortedatum = familie_werkblad.cell(rijnummer, 3).value
        # Voeg het ingelezen familielid toe aan de lijst.
        familielid = Familielid(naam=naam, familieverband=familieverband, geboortedatum=geboortedatum)
        familieleden.append(familielid)
        # Zet de leeftijd in kolom 4.
        leeftijd_cel = familie_werkblad.cell(rijnummer, 4)
        leeftijd_cel.value = familielid.leeftijd()

    # Bewaar de wijzigingen in hetzelfde bestand.
    if len(familieleden) > 0:
        familie_werkblad['D1'].value = 'Leeftijd'
        werkboek.save("bestanden/Familiebestand.xlsx")

    # Laat zien uit wie de familie bestaat.
    familie = 'geen familie'
    for rijnummer, familielid in enumerate(familieleden):
        # Voeg een scheidingstekst toe (niets, "en" of een komma).
        if rijnummer == 0:
            familie = ''
        else:
            if rijnummer == len(familieleden) - 1:
                familie = familie + ' en '
            else:
                familie = familie + ', '
        # Voeg de beschrijving van het familielid toe.
        familie = familie + str(familielid)

    print(f'Ik heb {familie}.')

    # Leg e.e.a. vast in de Postgres-database.
    open_database()

    geef_chocoladeletters(familieleden)
